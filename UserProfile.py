from openpyxl import load_workbook
import sys
import pickle


def save_obj(obj, name):
    with open('userdb/'+ name + '.pkl', 'wb') as f:
        pickle.dump(obj, f, pickle.HIGHEST_PROTOCOL)
def load_obj(name):
    with open('userdb/' + name + '.pkl', 'rb') as f:
        return pickle.load(f)

update = input("Would you like to update user profile? (y/n) ")
if update == 'y':
	"""Update user profile """ 
	wb = load_workbook(filename = 'registrations/iLibrary_registrations_28jul2016.xlsx',read_only=True)
	ws = wb['Sheet1']

	profile_dict = dict()
	row_count = ws.max_row-1
	for idx,row in enumerate(ws.rows):
		print("Updating registration file {0:.3f}% \r".format((idx/row_count)*100), end='')
		sys.stdout.flush()
		if profile_dict.get(row[1].value) is None:
			#{registered id : (name, email, country, cu number)}
			profile_dict[row[1].value] = (row[0].value, 
											row[3].value,
											row[9].value,
											row[10].value,)

	save_obj(profile_dict,'profile_dict')
	""" update ENDS HERE""" 

""" INTEGRATE TO THE USERDB """

u_d = load_obj('user_dict')
p_d = load_obj('profile_dict')

for idx, (u, d) in enumerate(u_d.items()):
	print("Updating userdb {0:.3f}% \r".format(((idx+1)/len(u_d))*100), end='')
	sys.stdout.flush()
	info = p_d.get(u)
	if info is not None:
		#no cu_number
		if info[3] is None:
			continue 
		d.identity_dict['user_name'] = info[0]
		d.identity_dict['email'] = info[1]
		d.identity_dict['geo'] = info[2]
		d.identity_dict['CU_number'] = info[3]

save_obj(u_d,'user_dict')
	


